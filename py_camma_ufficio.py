import openpyxl as xl
#camma cilindrica di un dato diametro es:720

wb1 = xl.load_workbook(filename="Cartel_1.xlsx")  # carico il file excel
sheet2 = wb1.create_sheet("Sheet2")
foglio1 = wb1.worksheets[0] #wb1.active attivo il primo sheet
foglio2 = wb1.worksheets[1]
hop = foglio1.max_row #ottengo il numero di righe

for i in range (1,hop+1):#col3_value = folgio1.value#creo la prima colonna con 360 - la terza colonna del foglio 1
    foglio2.cell(row=i, column=1).value = 360 #- foglio1.cell(row=i, column=3).value

for i in range (2,hop+1): #colonna 2 importo i valori come float così ottengo le virgole
    angolo=float(foglio1.cell(row=i, column=1).value)
    foglio2.cell(row=i-1, column=2).value = angolo * (-1) #moltiplico i valori per -1 per avere gli angoli negativi

for i in range (2,hop+1): #importo i valori della terza colonna come stringhe dato che non devo fare operazioni numeriche
    alzata=str(foglio1.cell(row=i, column=2).value)
    foglio2.cell(row=i-1, column=3).value = alzata

#sostituisco i punti con le stringhe, in questo caso solo colonna angolo
for i in range (1,hop+1):
    ang=str(foglio2.cell(row=i, column=2).value)
    foglio2.cell(row=i, column=2).value=ang
    #foglio2.cell(row=i, column=2).number_format = ang * (-1)


#creo il file txt
#seguo https://thispointer.com/how-to-append-text-or-lines-to-a-file-in-python/
for i in range (1, hop+1):
    out_text= open('camma.ibl','a')
    data1=str(foglio2.cell(row=i, column=1).value)+ str(' ')
    out_text.write(data1)
    data2=foglio2.cell(row=i, column=2).value + str(' ')
    out_text.write(data2)
    data3=foglio2.cell(row=i, column=3).value
    out_text.write(data3)
    out_text.write("\n")


#_____________________________________
#file1=open("file1.txt","w")
#file1.write(prima_colonna)

#ciclo for con i due indici i, j
#per righe e colonne per copiare le celle in foglio2
#for i in range (1,hop+1):
 #   for j in range (2,3):
 #       c = foglio1.cell(row = i, column = j)
 #       foglio2.cell(row=i, column = j+1).value=c.value

#foglio1.number_format = 'Comma'
wb1.save("Cartel1.xlsx")

#open('file1.txt', 'w') #creo il file .txt
#hop_range=range(1,hop)



# sheet_obj.max_row
# numero massimo di righe
#for value in sheet.iter_rows(min_row=1,
#                             max_row=2,
#                             min_col=1,
#                             max_col=3,
#                             values_only=True):
#    print(value)
#nome = workbook.sheetnames
#print(nome)
#prima_colonna=[]
#stampo tutti i valori della prima colonna
#for bss in sheet.iter_rows(min_row=1,
#                        max_col=1,
#                        values_only=True):
#    prima_colonna=bss
#file1=open("file1.txt","w")
#file1.write(prima_colonna)
#file1.close()
#sap = value[3]




#writing to a file
#with open('file1.txt', 'w') as f:
#    for item in prima_colonna:
#        f.write("%s\n" % item)


#return max row number in sheet_obj
#print(hop)

#my_array_r = []
#for value in sheet.iter_rows(min_row=1,
#                        max_col=1,
#                        values_only=True):

